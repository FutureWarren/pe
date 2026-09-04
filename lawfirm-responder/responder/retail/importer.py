"""把门店导出的库存表读进来——并且**先告诉他们这张表有什么问题**。

## 为什么要先有校验报告，再有导入

酷机时代的库存在一个后台小程序里。第一期最现实的路径不是对接接口，
是**每天从小程序后台导出一张表**丢给我们。而现实里那张表长什么样，
写代码的人事先永远猜不到：表头叫「售价」不叫「价格」、价格写成「￥6,499元」、
库存写成「城关3台/七里河1台」、更新时间那一列干脆空着。

如果导入器只是"能读的读、读不了的跳过"，结果就是一张残缺的表静静地进了系统，
然后 AI 因为查不到而把每个问价的客户都转人工——**功能看着正常，
实际上一条都没自动化**。这正是这套系统反复踩过的那类失败。

所以这一层的产出**首先是一份给人看的报告**：这张表有多少行能用、
哪几列没认出来、哪几行价格是空的、更新时间是不是过期了。
门店拿着报告改一遍表，比我们猜十次都管用。

## 表头认得越宽越好

同一个意思，不同后台导出来的叫法能有五六种。认不出来的代价是整列作废，
而多写几个别名的代价是零。所以别名表写宽，且大小写、空格、全半角都归一。
"""

import csv
import io
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from responder.retail.catalog import Catalog, Sku, _parse_stock

# 表头别名。**宁可多写，不可漏写**——漏一个别名，那一整列就作废了。
ALIASES: dict[str, tuple[str, ...]] = {
    "model": ("型号", "机型", "商品", "商品名称", "品名", "产品", "产品名称",
              "名称", "货品", "货品名称", "model", "name", "product"),
    "spec": ("配置", "规格", "版本", "内存", "容量", "存储", "参数",
             "spec", "config", "capacity"),
    "color": ("颜色", "配色", "色号", "color"),
    "price": ("价格", "售价", "零售价", "标价", "单价", "现价", "到手价",
              "销售价", "price", "amount"),
    "stock": ("库存", "数量", "台数", "可售", "可售数量", "现货", "剩余",
              "stock", "qty", "quantity", "inventory"),
    "promo": ("活动", "优惠", "促销", "备注", "说明", "活动说明", "promo", "note"),
    "updated_at": ("更新时间", "更新日期", "日期", "时间", "同步时间",
                   "updated", "updated_at", "date"),
}

_HEADER_LOOKUP: dict[str, str] = {}
for _field, _names in ALIASES.items():
    for _n in _names:
        _HEADER_LOOKUP[_n] = _field


def _norm_header(h: str) -> str:
    """归一化表头：去空格、去括号内容、全角转半角、转小写。

    「售价(元)」「售 价」「售价（含税）」都该认成同一列。
    """
    s = (h or "").strip()
    s = re.sub(r"[（(].*?[)）]", "", s)          # 去掉括号里的补充说明
    s = re.sub(r"[\s　_\-*]+", "", s)
    return s.lower()


def map_headers(headers: list[str]) -> tuple[dict[int, str], list[str]]:
    """把实际表头映射到我们认识的字段。返回 (列号→字段, 没认出来的表头)。"""
    mapping: dict[int, str] = {}
    unknown: list[str] = []
    for i, h in enumerate(headers):
        key = _HEADER_LOOKUP.get(_norm_header(h))
        if key and key not in mapping.values():
            mapping[i] = key
        elif h and h.strip():
            unknown.append(h.strip())
    return mapping, unknown


_MONEY_CHARS = re.compile(r"[^\d.]")


def parse_price(raw: str) -> int | None:
    """「￥6,499元」「6499.00」「6499 元」→ 6499；认不出返回 None。

    **认不出就是 None，绝不猜 0**——0 会被当成一个真实的价格报给客户。
    """
    s = _MONEY_CHARS.sub("", (raw or "").strip())
    if not s:
        return None
    try:
        v = float(s)
    except ValueError:
        return None
    return int(v) if v > 0 else None


def parse_when(raw: str, *, fallback: datetime | None = None) -> str:
    """把各种写法的日期归一成 ISO。认不出时用文件的修改时间兜底。

    为什么要兜底而不是留空：`catalog` 把「没有更新时间」视同过期、一律不报价。
    门店导出时漏掉这一列是常事，真让整张表因此失效，第一期就跑不起来了。
    **兜底用的是文件修改时间——它至少是真实的**，不是我们编的。
    """
    s = (raw or "").strip()
    if s:
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
                    "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d",
                    "%Y.%m.%d", "%m/%d/%Y", "%Y%m%d"):
            try:
                return datetime.strptime(s, fmt).isoformat()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(s).isoformat()
        except ValueError:
            pass
    return (fallback or datetime.now()).isoformat()


@dataclass
class Report:
    """给门店看的那份报告。**它比导入本身更重要。**"""

    total_rows: int = 0
    usable: int = 0
    mapped: dict[str, str] = field(default_factory=dict)   # 字段 → 实际表头
    unknown_headers: list[str] = field(default_factory=list)
    missing_fields: list[str] = field(default_factory=list)
    no_price: int = 0
    no_stock: int = 0
    borrowed_time: int = 0        # 更新时间是兜底填的，不是表里带的
    problems: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return self.usable > 0 and "model" not in self.missing_fields

    def to_text(self) -> str:
        """写成一段人话——收件人是门店的店长，不是工程师。"""
        lines = [f"读到 {self.total_rows} 行，其中 {self.usable} 行可用。"]
        if self.mapped:
            got = "、".join(f"{v}→{k}" for k, v in self.mapped.items())
            lines.append(f"认出来的列：{got}")
        if self.missing_fields:
            zh = {"model": "型号", "price": "价格", "stock": "库存"}
            miss = "、".join(zh.get(f, f) for f in self.missing_fields)
            lines.append(f"⚠️ 这几列没找到：{miss}")
        if self.unknown_headers:
            lines.append(f"没用上的列：{'、'.join(self.unknown_headers[:8])}")
        if self.no_price:
            lines.append(f"⚠️ {self.no_price} 行没有价格——这些机型 AI 不会报价，会转人工。")
        if self.no_stock:
            lines.append(f"提示：{self.no_stock} 行没有库存数，问「有没有货」时会转人工。")
        if self.borrowed_time:
            lines.append(
                f"提示：{self.borrowed_time} 行没写更新时间，已按文件的修改时间算。"
                f"**最好在导出时带上这一列**，否则隔天的表会被判成过期而停止报价。"
            )
        for p in self.problems:
            lines.append(f"⚠️ {p}")
        if not self.ok:
            lines.append("❌ 这张表还不能用，请按上面的提示改一版。")
        return "\n".join(lines)


def load(
    path: str | Path, *, max_age_hours: float = 24.0,
) -> tuple[Catalog, Report]:
    """读一张库存表（.csv / .xlsx），返回 (目录, 报告)。

    读不了的行不会静默丢掉——它们全部计进报告里，让人看得见。
    """
    p = Path(path)
    rows = _read_rows(p)
    rep = Report()
    if not rows:
        rep.problems.append("文件是空的，或者第一个工作表没有数据。")
        return Catalog([], max_age_hours=max_age_hours), rep

    headers, body = rows[0], rows[1:]
    mapping, unknown = map_headers(headers)
    rep.unknown_headers = unknown
    rep.mapped = {v: headers[k] for k, v in mapping.items()}
    for need in ("model", "price", "stock"):
        if need not in mapping.values():
            rep.missing_fields.append(need)

    file_mtime = datetime.fromtimestamp(p.stat().st_mtime) if p.exists() else None

    skus: list[Sku] = []
    for raw in body:
        rep.total_rows += 1
        g = {mapping[i]: (raw[i] if i < len(raw) else "")
             for i in mapping if i < len(raw)}
        model = (g.get("model") or "").strip()
        if not model:
            continue  # 没型号的行没有意义（多半是小计行或空行）
        price = parse_price(g.get("price", ""))
        stock = _parse_stock(g.get("stock", ""))
        had_time = bool((g.get("updated_at") or "").strip())
        if not had_time:
            rep.borrowed_time += 1
        if price is None:
            rep.no_price += 1
        if not stock:
            rep.no_stock += 1
        skus.append(Sku(
            model=model,
            spec=(g.get("spec") or "").strip(),
            color=(g.get("color") or "").strip(),
            price=price,
            stock=stock,
            promo=(g.get("promo") or "").strip(),
            updated_at=parse_when(g.get("updated_at", ""), fallback=file_mtime),
        ))
        rep.usable += 1

    return Catalog(skus, max_age_hours=max_age_hours), rep


def _read_rows(p: Path) -> list[list[str]]:
    """csv / xlsx 都读成一样的二维字符串表。"""
    suffix = p.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError:  # pragma: no cover - 依赖已在 pyproject 里
            raise ValueError("读 Excel 需要 openpyxl，请先 pip install openpyxl") from None
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        out = [[("" if c is None else str(c)) for c in row]
               for row in ws.iter_rows(values_only=True)]
        wb.close()
        return [r for r in out if any(x.strip() for x in r)]
    text = p.read_text(encoding="utf-8-sig", errors="replace")
    dialect = csv.excel
    if "\t" in text.split("\n")[0]:
        dialect = csv.excel_tab
    rows = list(csv.reader(io.StringIO(text), dialect))
    return [r for r in rows if any((x or "").strip() for x in r)]
