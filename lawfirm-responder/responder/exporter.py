"""客户档案导出：把线索整理成管理员能直接用的 Excel。

## 为什么要有这个

控制台是给「在系统里干活的人」用的：律师看自己的单、点已联系。但所主任要的
是另一样东西——**一份能打印、能转发、能在会上过一遍的表**。让他为了看昨天
进了多少客户去点开一个网页、翻一个列表，这件事他不会每天做。

所以导出不是「附加功能」，它是管理员这个角色的主视图。

## 列的取舍

只放管理员会用来做决定的列。AI 到底回了什么**不进表**——律所方原话是
「不想看到那么多 AI 对话，那么乱」。要看原文有深链，一列就够。

来源、案由、优先级、负责律师、状态、跟进时长——这六列回答的是同一个问题：
**这些人有没有被好好接住。**
"""

import io
import logging
from datetime import datetime

from responder.config import Settings, get_settings
from responder.store.db import Store

logger = logging.getLogger(__name__)

# 表头与取值函数。顺序即列序，改这里就改了导出格式。
_COLUMNS: list[tuple[str, str]] = [
    ("进线时间", "created_at"),
    ("客户", "_name"),
    ("联系方式", "contact"),
    ("来源", "_source"),
    ("案由", "case_type"),
    ("优先级", "priority"),
    ("评分", "score"),
    ("客户诉求", "summary"),
    ("关键信息", "_facts"),
    ("负责律师", "_lawyer"),
    ("状态", "_status"),
    ("跟进时长(小时)", "_hours"),
    ("备注", "notes"),
    ("完整对话", "_link"),
]

_STATUS_CN = {
    "new": "待跟进",
    "contacted": "已联系",
    "converted": "已成交",
    "invalid": "无效",
}


def _when(iso: str) -> str:
    """ISO 时间戳 → 「08-06 15:47」。

    `2026-08-06T07:47:34.340981` 这种东西不该出现在给所主任看的表里：
    它不是给人读的，而且列宽会被那串微秒撑成两倍。
    """
    try:
        return datetime.fromisoformat(iso).strftime("%m-%d %H:%M")
    except (TypeError, ValueError):
        return iso or ""


# 渠道标识 → 给人看的名字。表里写「meituan」没人看得懂，
# 而这一列是管理员判断「钱该往哪加」的唯一依据。
_CHANNEL_CN = {
    "meituan": "美团", "dianping": "点评", "xhs": "小红书",
    "douyin": "抖音", "wechat": "微信", "baidu": "百度",
    "shipinhao": "视频号", "web": "官网", "offline": "线下",
}


def _source(group_id: str, group=None) -> str:
    """线索从哪来。

    优先用会话档案上记的渠道：所有从微信客服进来的人在库里长得一模一样，
    光看 group_id 只能答出「微信客服」——而视频号来的、官网来的、名片扫的
    是完全不同的三笔生意，分不开就说不清哪个渠道该加钱。
    """
    channel = getattr(group, "ext_channel", "") if group else ""
    if channel:
        label = _CHANNEL_CN.get(channel, channel)
        return label
    if group_id.startswith("kf:"):
        return "微信客服"
    if group_id.startswith(("dy:", "dyim:")):
        return "抖音"
    if group_id.startswith("ch:"):
        # 档案缺渠道字段（老数据）时从 group_id 里兜一层
        part = group_id.split(":")[1] if group_id.count(":") >= 2 else ""
        return _CHANNEL_CN.get(part, part or "外部渠道")
    return "群聊"


def _hours_to_contact(row: dict) -> str:
    """派单到首次跟进的间隔。线索的价值随时间塌得很快，这一列是给人看的秒表。"""
    a, u = row.get("assigned_at"), row.get("updated_at")
    if not a or not u or row.get("status") == "new":
        return ""
    try:
        delta = datetime.fromisoformat(u) - datetime.fromisoformat(a)
    except (TypeError, ValueError):
        return ""
    return str(round(delta.total_seconds() / 3600, 1))


def build_rows(
    store: Store, leads: list[dict], settings: Settings | None = None
) -> list[list[str]]:
    """线索行 → 表格行（含表头）。纯函数，便于测试与复用到 CSV。"""
    import json

    settings = settings or get_settings()
    base = settings.public_base_url.rstrip("/")
    names = {law["userid"]: (law.get("name") or law["userid"])
             for law in store.list_lawyers()}
    out = [[title for title, _ in _COLUMNS]]
    for row in leads:
        gid = row.get("group_id", "")
        group = store.get_group(gid)
        try:
            facts = json.loads(row.get("key_facts") or "[]")
        except (ValueError, TypeError):
            facts = []
        computed = {
            "created_at": _when(row.get("created_at", "")),
            "_name": (group.name if group else "") or row.get("contact", "") or "—",
            "_source": _source(gid, group),
            "_facts": "；".join(str(f) for f in facts),
            "_lawyer": names.get(row.get("assigned_userid", ""), "") or "未派单",
            "_status": _STATUS_CN.get(row.get("status", ""), row.get("status", "")),
            "_hours": _hours_to_contact(row),
            # 深链而不是把对话糊进单元格：管理员不想看 AI 对话，
            # 但偶尔需要点进去看一眼。一列链接兼顾两者。
            # 用 `/g/{id}`，不用 `/ui#g={id}`：后者在企业微信里点开是 404
            # （企微把 `#` 转义成 `%23`），而这张表是会被转发进企微群的。
            "_link": f"{base}/g/{gid}" if base else "",
        }
        out.append([
            _safe_cell(str(computed.get(key, row.get(key, "")) or ""))
            for _, key in _COLUMNS
        ])
    return out


# Excel 把 = + - @ 开头的单元格当公式执行。客户诉求是**客户自己打的字**，
# 一句「=1+1」在表里会变成 2，而 `=cmd|...` 这类在某些环境下能拉起外部程序。
# 律所会把这张表转发给合伙人、发进微信群——我们不能把一个可执行的东西发出去。
# 前面加一个单引号即可让 Excel 按纯文本处理，肉眼几乎看不出来。
_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(value: str) -> str:
    return "'" + value if value[:1] in _FORMULA_LEAD else value


def to_xlsx(rows: list[list[str]], title: str = "客户档案") -> bytes:
    """写成 Excel。openpyxl 缺失时抛 ValueError，由上层回落 CSV。"""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise ValueError("服务器未安装 openpyxl，请改用 CSV 导出") from e

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:30]
    for r in rows:
        ws.append(r)

    # 表头样式 + 冻结首行：几百行的表不冻结首行，翻到一半就不知道哪列是哪列
    head_fill = PatternFill("solid", fgColor="1F3A5F")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # 列宽按内容估：中文按两个字符算，否则「客户诉求」那列永远是一条窄缝
    for i, _ in enumerate(rows[0] if rows else [], start=1):
        width = 0
        for row in rows[:200]:  # 只看前 200 行，几百行全扫没必要
            if i <= len(row):
                text = row[i - 1]
                width = max(width, sum(2 if ord(c) > 127 else 1 for c in text))
        ws.column_dimensions[get_column_letter(i)].width = min(max(width + 2, 8), 48)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv(rows: list[list[str]]) -> bytes:
    """CSV 兜底。用 GBK 兼容的 utf-8-sig：不带 BOM 的话 Excel 打开就是乱码。"""
    import csv

    buf = io.StringIO()
    csv.writer(buf).writerows(rows)
    return buf.getvalue().encode("utf-8-sig")
