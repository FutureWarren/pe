"""Analyst-ready workbook writer.

Produces exactly three sheets — Model_Input, Exceptions, Source_Map — from
the canonical ModelInputBundle. No trace tables, no debug dumps on the
analyst-facing sheets.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.export.safe_cell import sanitize_cell_text
from app.models.canonical import (
    DERIVED_METRIC_FORMULAS,
    EXCEL_NUMBER_FORMAT,
    ExceptionRow,
    FinalMetricRecord,
    METRIC_DISPLAY,
    METRIC_ORDER,
    MetricUnit,
    ModelInputBundle,
)

MODEL_INPUT_SHEET = "Model_Input"
EXCEPTIONS_SHEET = "Exceptions"
SOURCE_MAP_SHEET = "Source_Map"

_HEADER_FILL = PatternFill(start_color="FF1F3950", end_color="FF1F3950", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF", name="Calibri", size=11)
_SECTION_FONT = Font(bold=True, name="Calibri", size=11, color="FF1F3950")
_METRIC_FONT = Font(bold=True, name="Calibri", size=10)
_BODY_FONT = Font(name="Calibri", size=10)
_THIN_SIDE = Side(border_style="thin", color="FFDFE3EB")
_CELL_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)

_STATUS_COLOURS = {
    "Ready": PatternFill(start_color="FFDFF3E3", end_color="FFDFF3E3", fill_type="solid"),
    "Review": PatternFill(start_color="FFFCE8D5", end_color="FFFCE8D5", fill_type="solid"),
}
_CONFIDENCE_COLOURS = {
    "High": PatternFill(start_color="FFDFF3E3", end_color="FFDFF3E3", fill_type="solid"),
    "Medium": PatternFill(start_color="FFFFF5CF", end_color="FFFFF5CF", fill_type="solid"),
    "Low": PatternFill(start_color="FFF8D4D4", end_color="FFF8D4D4", fill_type="solid"),
}

_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def write_analyst_workbook(output_path: Path, bundle: ModelInputBundle) -> Path:
    """Write the three analyst-facing sheets to `output_path`."""

    workbook = Workbook()
    # First sheet created automatically — rename and use for Model_Input.
    model_input_sheet = workbook.active
    model_input_sheet.title = MODEL_INPUT_SHEET
    exceptions_sheet = workbook.create_sheet(title=EXCEPTIONS_SHEET)
    source_map_sheet = workbook.create_sheet(title=SOURCE_MAP_SHEET)

    _write_model_input(model_input_sheet, bundle)
    _write_exceptions(exceptions_sheet, bundle.exceptions)
    _write_source_map(source_map_sheet, bundle.metrics)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Model_Input (wide table)
# ---------------------------------------------------------------------------


def _write_model_input(sheet: Worksheet, bundle: ModelInputBundle) -> None:
    period_labels = bundle.period_order
    period_keys = bundle.period_keys
    metric_rows = _metric_rows_index(bundle.metrics)

    static_headers = ["Metric", *period_labels, "Unit", "Confidence", "Validation", "Status", "Notes"]
    for column_index, header in enumerate(static_headers, start=1):
        cell = sheet.cell(row=1, column=column_index, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _CELL_BORDER

    first_period_column = 2
    last_period_column = 1 + len(period_labels)

    # Write metric rows.
    for row_offset, metric_key in enumerate(METRIC_ORDER):
        row_index = 2 + row_offset
        sheet.cell(row=row_index, column=1, value=METRIC_DISPLAY[metric_key]).font = _METRIC_FONT
        sheet.cell(row=row_index, column=1).alignment = _LEFT
        sheet.cell(row=row_index, column=1).border = _CELL_BORDER

        period_records = metric_rows.get(metric_key, {})
        representative: Optional[FinalMetricRecord] = None
        if period_records:
            representative = next(iter(period_records.values()))

        summary = _summarize_metric_row(list(period_records.values()))
        unit_label = _unit_display(summary["unit"] if representative else None, metric_key)
        confidence = summary["confidence"]
        validation = summary["validation"]
        status = summary["status"]
        note = summary["note"]

        # Period columns.
        for period_index, period_key in enumerate(period_keys):
            column_letter = get_column_letter(first_period_column + period_index)
            column_index = first_period_column + period_index
            record = period_records.get(period_key)
            cell = sheet.cell(row=row_index, column=column_index)

            if record is None:
                cell.value = None
            elif metric_key in DERIVED_METRIC_FORMULAS:
                formula = _build_excel_formula(
                    metric_key,
                    row_index,
                    column_letter,
                )
                if formula is not None:
                    cell.value = formula
                else:
                    cell.value = record.final_value
            else:
                cell.value = record.final_value

            cell.number_format = _excel_number_format(record.unit if record else (representative.unit if representative else None))
            cell.font = _BODY_FONT
            cell.alignment = _CENTER
            cell.border = _CELL_BORDER
            if record is not None and record.status == "Review":
                cell.fill = _STATUS_COLOURS["Review"]

        # Metadata columns.
        metadata_columns = [
            (last_period_column + 1, unit_label, _LEFT, None),
            (last_period_column + 2, confidence, _CENTER, _CONFIDENCE_COLOURS.get(confidence)),
            (last_period_column + 3, validation, _CENTER, None),
            (last_period_column + 4, status, _CENTER, _STATUS_COLOURS.get(status)),
            (last_period_column + 5, note or "", _LEFT, None),
        ]
        for column_index, value, alignment, fill in metadata_columns:
            cell = sheet.cell(row=row_index, column=column_index, value=sanitize_cell_text(value))
            cell.font = _BODY_FONT
            cell.alignment = alignment
            cell.border = _CELL_BORDER
            if fill is not None:
                cell.fill = fill

    # Layout niceties.
    sheet.freeze_panes = "B2"
    sheet.column_dimensions["A"].width = 26
    for period_index in range(len(period_labels)):
        sheet.column_dimensions[get_column_letter(first_period_column + period_index)].width = 14
    for offset, width in enumerate((10, 12, 14, 10, 36), start=1):
        sheet.column_dimensions[get_column_letter(last_period_column + offset)].width = width


def _metric_rows_index(
    metrics: list[FinalMetricRecord],
) -> dict[str, dict[str, FinalMetricRecord]]:
    """Arrange metrics into {metric_key: {period_key: record}}."""

    index: dict[str, dict[str, FinalMetricRecord]] = {}
    for metric in metrics:
        index.setdefault(metric.metric_key, {})[metric.period_key] = metric
    return index


def _summarize_metric_row(records: list[FinalMetricRecord]) -> dict[str, Optional[str | MetricUnit]]:
    if not records:
        return {
            "unit": None,
            "confidence": "",
            "validation": "",
            "status": "",
            "note": "",
        }

    confidence_rank = {"High": 0, "Medium": 1, "Low": 2}
    worst_confidence = max(records, key=lambda record: confidence_rank[record.confidence_level]).confidence_level

    if any(record.validation_result == "Mismatch" for record in records):
        validation = "Mismatch"
    elif any(record.validation_result == "Formula" for record in records):
        validation = "Formula"
    elif any(record.validation_result == "Matched" for record in records):
        validation = "Matched"
    else:
        validation = "Single-source"

    status = "Review" if any(record.status == "Review" for record in records) else "Ready"
    note = next((record.note for record in records if record.note), "") or ""
    unit = next((record.unit for record in records if record.unit is not None), None)

    return {
        "unit": unit,
        "confidence": worst_confidence,
        "validation": validation,
        "status": status,
        "note": note,
    }


def _build_excel_formula(
    metric_key: str,
    row_index: int,
    column_letter: str,
) -> Optional[str]:
    """Return an Excel formula string that references sibling metric rows."""

    metric_row = {key: 2 + index for index, key in enumerate(METRIC_ORDER)}

    def ref(key: str) -> str:
        return f"{column_letter}{metric_row[key]}"

    if metric_key == "gross_profit":
        return f"=IF(OR({ref('revenue')}=\"\",{ref('cogs')}=\"\"),\"\",{ref('revenue')}-{ref('cogs')})"
    if metric_key == "ebitda":
        return f"=IF(OR({ref('gross_profit')}=\"\",{ref('operating_expenses')}=\"\"),\"\",{ref('gross_profit')}-{ref('operating_expenses')})"
    if metric_key == "gross_margin_pct":
        return f"=IF(OR({ref('revenue')}=\"\",{ref('revenue')}=0),\"\",{ref('gross_profit')}/{ref('revenue')})"
    if metric_key == "ebitda_margin_pct":
        return f"=IF(OR({ref('revenue')}=\"\",{ref('revenue')}=0),\"\",{ref('ebitda')}/{ref('revenue')})"
    return None


def _excel_number_format(unit: Optional[MetricUnit]) -> str:
    if unit is None:
        return "General"
    return EXCEL_NUMBER_FORMAT.get(unit, "General")


def _unit_display(unit: Optional[MetricUnit], metric_key: str) -> str:
    if unit == "%":
        return "%"
    if unit == "count":
        return "count"
    if unit == "USD_thousands":
        return "USD '000"
    return "USD"


# ---------------------------------------------------------------------------
# Exceptions sheet
# ---------------------------------------------------------------------------


def _write_exceptions(sheet: Worksheet, exceptions: list[ExceptionRow]) -> None:
    headers = ["Metric", "Period", "Issue", "System_View", "Suggested_Action", "Severity"]
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _CELL_BORDER

    if not exceptions:
        summary_cell = sheet.cell(row=2, column=1, value="No exceptions. All metrics passed validation.")
        summary_cell.font = _SECTION_FONT
        sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
        return

    severity_order = {"Critical": 0, "Review": 1, "Info": 2}
    sorted_exceptions = sorted(
        exceptions, key=lambda row: (severity_order.get(row.severity, 99), row.metric, row.period)
    )
    for row_offset, exception in enumerate(sorted_exceptions, start=2):
        row_values = [
            exception.metric,
            exception.period,
            exception.issue,
            exception.system_view,
            exception.suggested_action,
            exception.severity,
        ]
        for column_index, value in enumerate(row_values, start=1):
            cell = sheet.cell(row=row_offset, column=column_index, value=sanitize_cell_text(value))
            cell.font = _BODY_FONT
            cell.alignment = _LEFT
            cell.border = _CELL_BORDER

        severity_fill = {
            "Critical": PatternFill(start_color="FFF8D4D4", end_color="FFF8D4D4", fill_type="solid"),
            "Review": PatternFill(start_color="FFFCE8D5", end_color="FFFCE8D5", fill_type="solid"),
            "Info": PatternFill(start_color="FFE0ECF9", end_color="FFE0ECF9", fill_type="solid"),
        }.get(exception.severity)
        if severity_fill is not None:
            sheet.cell(row=row_offset, column=6).fill = severity_fill

    for column_index, width in enumerate((22, 12, 32, 48, 48, 12), start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    sheet.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Source_Map sheet
# ---------------------------------------------------------------------------


def _write_source_map(sheet: Worksheet, metrics: list[FinalMetricRecord]) -> None:
    headers = [
        "Metric",
        "Period",
        "Final_Value",
        "Primary_Source_File",
        "Primary_Source_Tab",
        "Primary_Source_Range",
        "Backup_Sources",
        "Direct_or_Derived",
        "Derivation_Formula",
        "Validation_Result",
        "Confidence_Reason",
        "Source_Priority_Reason",
    ]
    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _CELL_BORDER

    # Stable order: metric order, then period order.
    metric_rank = {key: index for index, key in enumerate(METRIC_ORDER)}
    ordered = sorted(
        metrics, key=lambda m: (metric_rank.get(m.metric_key, 999), m.period_order)
    )

    for row_offset, metric in enumerate(ordered, start=2):
        primary = metric.selected_source
        backups_text = _format_backups(metric)
        values = [
            metric.metric_name,
            metric.period,
            _format_final_value(metric),
            primary.file if primary else "",
            primary.tab if primary and primary.tab else "",
            primary.range if primary and primary.range else "",
            backups_text,
            "Direct" if metric.direct_or_derived == "direct" else "Derived",
            metric.derivation_formula or "",
            metric.validation_result,
            metric.confidence_reason,
            metric.source_priority_reason or "",
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_offset, column=column_index, value=sanitize_cell_text(value))
            cell.font = _BODY_FONT
            cell.alignment = _LEFT
            cell.border = _CELL_BORDER

    for column_index, width in enumerate(
        (22, 12, 14, 32, 22, 18, 32, 14, 28, 16, 48, 48),
        start=1,
    ):
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    sheet.freeze_panes = "A2"


def _format_backups(metric: FinalMetricRecord) -> str:
    if not metric.backup_sources:
        return ""
    parts = []
    for citation in metric.backup_sources:
        pieces = [citation.file]
        if citation.tab:
            pieces.append(citation.tab)
        if citation.range:
            pieces.append(citation.range)
        parts.append(" → ".join(p for p in pieces if p))
    return "\n".join(parts)


def _format_final_value(metric: FinalMetricRecord) -> str:
    if metric.final_value is None:
        return ""
    if metric.unit == "%":
        return f"{metric.final_value * 100:.2f}%"
    if metric.unit == "count":
        return f"{metric.final_value:,.0f}"
    return f"{metric.final_value:,.0f}"
