"""Deterministic workbook writing for the pilot pipeline."""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.export.safe_cell import sanitize_cell_text
from app.map.cell_rules import (
    P_AND_L_SHEET_NAME,
    SOURCE_MAP_SHEET_NAME,
    SOURCE_REGISTRY_SHEET_NAME,
    VALIDATION_SHEET_NAME,
)
from app.models.mapping import SourceMapEntry, WorkbookCellBinding
from app.models.source import SourceManifest
from app.models.validation import ValidationReport


def write_scaffold_workbook(
    output_path: Path,
    template_workbook_path: Optional[Path],
    manifest: SourceManifest,
    cell_bindings: list[WorkbookCellBinding],
    source_map_entries: list[SourceMapEntry],
    validation_report: ValidationReport,
) -> Path:
    """Write a deterministic workbook with P&L, source traceability, and validation tabs."""

    workbook = _load_or_create_workbook(template_workbook_path)
    pnl_sheet = _ensure_sheet(workbook, P_AND_L_SHEET_NAME)
    source_map_sheet = _ensure_sheet(workbook, SOURCE_MAP_SHEET_NAME)
    source_registry_sheet = _ensure_sheet(workbook, SOURCE_REGISTRY_SHEET_NAME)
    validation_sheet = _ensure_sheet(workbook, VALIDATION_SHEET_NAME)

    _apply_bindings(workbook, cell_bindings)
    _write_source_registry(source_registry_sheet, manifest)
    _write_source_map(source_map_sheet, source_map_entries)
    _write_validation_sheet(validation_sheet, validation_report)
    _style_pnl_sheet(pnl_sheet)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _load_or_create_workbook(template_workbook_path: Optional[Path]) -> Workbook:
    """Load a template workbook when present or create a new workbook."""

    if template_workbook_path and template_workbook_path.exists():
        return load_workbook(template_workbook_path)
    workbook = Workbook()
    workbook.active.title = P_AND_L_SHEET_NAME
    return workbook


def _ensure_sheet(workbook: Workbook, sheet_name: str) -> Worksheet:
    """Return a worksheet, creating it when needed."""

    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    return workbook.create_sheet(title=sheet_name)


def _apply_bindings(workbook: Workbook, bindings: list[WorkbookCellBinding]) -> None:
    """Apply deterministic cell bindings to the workbook."""

    for binding in bindings:
        sheet = workbook[binding.sheet_name]
        cell = sheet[binding.cell]
        if binding.formula is not None:
            cell.value = binding.formula
        else:
            cell.value = sanitize_cell_text(binding.value)
        cell.number_format = binding.number_format
        if binding.comment:
            cell.comment = Comment(binding.comment, "Angelic Pilot")
        if binding.hyperlink:
            cell.hyperlink = binding.hyperlink


def _write_source_registry(sheet: Worksheet, manifest: SourceManifest) -> None:
    """Write the indexed source registry to a workbook tab."""

    headers = [
        "source_id",
        "file_name",
        "rel_path",
        "file_type",
        "modified_timestamp",
        "content_fingerprint",
        "document_role",
        "priority_rank",
    ]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=index, value=header).font = Font(bold=True)

    for row_index, document in enumerate(manifest.documents, start=2):
        sheet.cell(row=row_index, column=1, value=document.source_id)
        sheet.cell(row=row_index, column=2, value=sanitize_cell_text(document.file_name))
        sheet.cell(row=row_index, column=3, value=sanitize_cell_text(document.rel_path))
        sheet.cell(row=row_index, column=4, value=document.file_type)
        sheet.cell(row=row_index, column=5, value=document.modified_timestamp)
        sheet.cell(row=row_index, column=6, value=document.content_fingerprint)
        sheet.cell(row=row_index, column=7, value=document.document_role)
        sheet.cell(row=row_index, column=8, value=document.priority_rank)


def _write_source_map(sheet: Worksheet, entries: list[SourceMapEntry]) -> None:
    """Write the source map tab for populated cells."""

    headers = ["sheet", "cell", "line_item", "period_key", "value", "source_ids", "locators", "quotes"]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=index, value=header).font = Font(bold=True)

    for row_index, entry in enumerate(entries, start=2):
        sheet.cell(row=row_index, column=1, value=entry.sheet_name)
        sheet.cell(row=row_index, column=2, value=entry.cell)
        sheet.cell(row=row_index, column=3, value=sanitize_cell_text(entry.line_item_code))
        sheet.cell(row=row_index, column=4, value=entry.period_key)
        sheet.cell(row=row_index, column=5, value=sanitize_cell_text(entry.value_display))
        sheet.cell(row=row_index, column=6, value=", ".join(entry.source_ids))
        sheet.cell(row=row_index, column=7, value=sanitize_cell_text(" | ".join(entry.locators)))
        sheet.cell(row=row_index, column=8, value=sanitize_cell_text(" | ".join(entry.quotes)))


def _write_validation_sheet(sheet: Worksheet, report: ValidationReport) -> None:
    """Write the validation report into the workbook."""

    headers = ["severity", "code", "message", "context"]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=index, value=header).font = Font(bold=True)

    for row_index, issue in enumerate(report.issues, start=2):
        sheet.cell(row=row_index, column=1, value=issue.severity)
        sheet.cell(row=row_index, column=2, value=issue.code)
        sheet.cell(row=row_index, column=3, value=issue.message)
        sheet.cell(row=row_index, column=4, value=str(issue.context))

    assumption_row = len(report.issues) + 4
    sheet.cell(row=assumption_row, column=1, value="assumptions").font = Font(bold=True)
    for index, assumption in enumerate(report.assumptions, start=assumption_row + 1):
        sheet.cell(row=index, column=1, value=assumption)


def _style_pnl_sheet(sheet: Worksheet) -> None:
    """Apply lightweight demo-friendly styling to the P&L sheet."""

    sheet.freeze_panes = "B5"
    sheet.column_dimensions["A"].width = 28
    for column in ("B", "C", "D", "E", "F", "G", "H", "I", "J", "K"):
        sheet.column_dimensions[column].width = 18
    for cell in ("A1", "A4"):
        sheet[cell].font = Font(bold=True)
