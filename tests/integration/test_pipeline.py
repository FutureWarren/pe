from pathlib import Path

from openpyxl import load_workbook

from app.models.run import RunRequest
from app.services.pipeline import run_pilot


def test_run_pilot_creates_workbook_and_artifacts(tmp_path: Path) -> None:
    data_room = tmp_path / "data_room"
    data_room.mkdir()
    (data_room / "demo_pnl.csv").write_text(
        "Metric,FY2023,FY2024\n"
        "Revenue,100,120\n"
        "COGS,40,45\n"
        "Operating Expenses,20,25\n",
        encoding="utf-8",
    )
    (data_room / "notes.txt").write_text(
        "Top customer concentration FY2024: 35%\nEmployee count FY2024: 110\n",
        encoding="utf-8",
    )

    output_root = tmp_path / "outputs"
    summary = run_pilot(
        RunRequest(
            data_room_dir=data_room,
            output_root=output_root,
            run_label="integration-demo",
        ),
    )

    assert summary.status == "completed"
    assert summary.extraction_backend == "deterministic"
    assert summary.run_label == "integration-demo"
    assert summary.validation_status in {"pass", "warning", "fail"}
    assert summary.document_count == 2
    assert (summary.output_dir / "source_registry.json").exists()
    assert (summary.output_dir / "extracted_pnl.json").exists()
    assert (summary.output_dir / "resolved_pnl.json").exists()
    assert (summary.output_dir / "validation_report.json").exists()
    assert summary.workbook_path.exists()

    workbook = load_workbook(summary.workbook_path)
    assert workbook.sheetnames == ["Model_Input", "Exceptions", "Source_Map"]

    model_input_sheet = workbook["Model_Input"]
    assert model_input_sheet["B1"].value == "FY2023"
    assert model_input_sheet["C1"].value == "FY2024"
    assert model_input_sheet["B2"].value == 100
    assert model_input_sheet["B3"].value == 40
    assert model_input_sheet["B4"].value == '=IF(OR(B2="",B3=""),"",B2-B3)'
    assert model_input_sheet["C4"].value == '=IF(OR(C2="",C3=""),"",C2-C3)'

    source_map_sheet = workbook["Source_Map"]
    assert source_map_sheet["A1"].value == "Metric"
    assert source_map_sheet["B1"].value == "Period"
